# HTTPライブラリ
import requests
# スクレイピングライブラリ
from bs4 import BeautifulSoup
# データ解析ライブラリ
import pandas as pd
# 時間計測ライブラリ
import time
# Excel操作用ライブラリ
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

# DataFrame用の配列変数を定義、初期化
data = []

# 検索対象の商品カテゴリー
category = "CBD"

# 最初の5ページを取得
for i in range(1, 6):
    url = f"https://www.amazon.co.jp/s?k={category}&page={i}"

    print(f"Accessing page {i}...")

    # ページの内容を取得
    res = requests.get(url)
    res.raise_for_status()

    # BeautifulSoupオブジェクトを作成
    soup = BeautifulSoup(res.text, "html.parser")

    print("Parsing page...")

    # 商品の情報を抽出
    items = soup.select('.s-result-item')

    for item in items:
        # 商品名
        name = item.select_one('.a-link-normal .a-text-normal')
        # レビュー数
        review_count = item.select_one('.a-size-small .a-link-normal')
        # レビュー評価
        rating = item.select_one('.a-icon-alt')

        if name and review_count and rating:
            data.append({
                "商品名": name.text.strip(),
                "レビュー数": review_count.text.strip(),
                "レート": rating.text.strip(),
            })

    # 2秒スリープ
    print("Sleeping for 2 second...")
    time.sleep(2)

# DataFrameを作成
df = pd.DataFrame(data)

# Excelに出力
print("Writing to Excel...")
df.to_excel('output.xlsx', index=False)

print("Adjusting column widths, text wrapping, and cell borders in Excel...")
wb = load_workbook('output.xlsx')
sheet = wb.active

# 商品名列の列幅を69に設定
col_idx = df.columns.get_loc("商品名") + 1  # pandasの列番号は0から始まるが、openpyxlの列番号は1から始まる
col_letter = get_column_letter(col_idx)
sheet.column_dimensions[col_letter].width = 69

# 商品名列の全てのセルに折り返しを適用
for row in sheet[col_letter]:
    row.alignment = Alignment(wrap_text=True)

# B列とC列の列幅を自動調整（カラム名も考慮）
for col_name in ["レビュー数", "レート"]:
    col_idx = df.columns.get_loc(col_name) + 1
    col_letter = get_column_letter(col_idx)
    max_length = 0
    column = sheet[col_letter]
    for cell in column:
        try: 
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        except:
            pass
    sheet.column_dimensions[col_letter].width = max_length

# 全てのセルに罫線を追加
thin = Side(border_style="thin")
for row in sheet.iter_rows():
    for cell in row:
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Save the changes
wb.save('output.xlsx')

print("Done!")
